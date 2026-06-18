#!/usr/bin/env python3
import argparse
import logging
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz

# ================================================================
# 1. Logging
# ================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("publisher_standardisation.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

# ================================================================
# 2. Configuration
# ================================================================
SIMILARITY_THRESHOLD_AUTO = 85
SIMILARITY_THRESHOLD_REVIEW = 65
AUTO_BOOST_THRESHOLD = 70
FREQ_RATIO_FOR_AUTO = 3

STOPWORDS = {
    "the", "of", "and", "for", "on", "at", "to", "in", "with", "without",
    "publishing", "press", "group", "inc", "ltd", "limited", "corporation",
    "gmbh", "ag", "verlag", "editions", "publications", "books", "house",
    "university", "universidade", "college", "institute", "school", "academy",
    "society", "association", "federation", "foundation", "centre", "center",
    "journal", "review", "research", "studies", "international", "national",
    "european", "american", "british", "federal", "state", "province",
    "regional", "local", "global", "world", "central", "union",
}

MAJOR_PUBLISHER_TOKENS = {
    "elsevier", "springer", "wiley", "taylor", "francis", "sage", "de gruyter",
    "oxford", "cambridge", "emerald", "inderscience", "mdpi", "plos",
    "frontiers", "bmc", "biomed central", "nature", "palgrave", "macmillan",
    "routledge", "informa", "healthcare",
}

# ================================================================
# 3. Hard-coded canonical mapping
# ================================================================
def apply_hardcoded_canonical(name):
    if pd.isna(name):
        return name

    name_lower = str(name).lower()

    if "taylor" in name_lower and "francis" in name_lower:
        return "Taylor & Francis"
    if "ieee" in name_lower or "institute of electrical and electronics engineers" in name_lower:
        return "Institute of Electrical and Electronics Engineers (IEEE)"
    if "biomed central" in name_lower or "bmc" in name_lower:
        return "BioMed Central"

    if "elsevier" in name_lower:
        return "Elsevier"
    if "springer" in name_lower:
        return "Springer"
    if "wiley" in name_lower:
        return "Wiley"
    if "sage" in name_lower:
        return "SAGE Publications"
    if "frontiers" in name_lower:
        return "Frontiers"
    if "plos" in name_lower:
        return "PLOS"
    if "mdpi" in name_lower:
        return "MDPI"
    if "de gruyter" in name_lower:
        return "De Gruyter"
    if "oxford university press" in name_lower or "oup" in name_lower:
        return "Oxford University Press"
    if "cambridge university press" in name_lower:
        return "Cambridge University Press"
    if "emerald" in name_lower:
        return "Emerald Publishing"
    if "inderscience" in name_lower:
        return "Inderscience"
    if "world scientific" in name_lower:
        return "World Scientific"

    return name

# ================================================================
# 4. Helpers
# ================================================================
def preprocess_name(name):
    if pd.isna(name):
        return []
    s = str(name).lower()
    s = s.replace("&", " and ")
    s = re.sub(r"[^\w\s]", " ", s)
    tokens = s.split()
    tokens = [t for t in tokens if t not in STOPWORDS]
    return tokens

def get_group_key(name):
    tokens = preprocess_name(name)
    return tokens[0] if tokens else str(name)[:3].upper()

def build_components(edges, num_items):
    parent = list(range(num_items))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[ry] = rx

    for i, j in edges:
        union(i, j)

    comp_map = {}
    for i in range(num_items):
        root = find(i)
        comp_map.setdefault(root, []).append(i)

    comp_id_for_item = {}
    for comp_id, members in enumerate(comp_map.values(), start=1):
        for idx in members:
            comp_id_for_item[idx] = comp_id
    return comp_id_for_item

def load_overrides(override_file):
    if not override_file or not Path(override_file).exists():
        return {}

    logger.info(f"Loading overrides from {override_file}")
    try:
        if str(override_file).lower().endswith(".csv"):
            ov = pd.read_csv(override_file)
        else:
            ov = pd.read_excel(override_file)
    except Exception as e:
        logger.error(f"Failed to read override file: {e}")
        return {}

    overrides = {}
    for _, row in ov.iterrows():
        name1 = str(row["publisher_1"]).strip()
        name2 = str(row["publisher_2"]).strip()
        action = str(row["action"]).strip().lower()

        if action == "keep":
            val = None
        elif action == "merge":
            val = "MERGE_AUTO_CANONICAL"
        else:
            val = action

        overrides[(name1, name2)] = val
        overrides[(name2, name1)] = val

    logger.info(f"Loaded {len(overrides) // 2} override pairs.")
    return overrides

def choose_component_canonical(variants, freq_counter):
    freq_best = max(variants, key=lambda v: (freq_counter[v], -len(str(v))))
    exact_major = [v for v in variants if apply_hardcoded_canonical(v) != v]
    if exact_major:
        mapped = apply_hardcoded_canonical(freq_best)
        if mapped != freq_best:
            return mapped
    return freq_best

# ================================================================
# 5. Main pipeline
# ================================================================
def process_publishers(input_file, output_file, review_file, cluster_file, override_file=None):
    logger.info("=" * 70)
    logger.info("Publisher standardisation pipeline")
    logger.info("=" * 70)

    input_path = Path(input_file)
    if not input_path.exists():
        logger.error(f"Input file not found: {input_file}")
        sys.exit(1)

    try:
        df = pd.read_excel(input_file, engine="openpyxl")
    except Exception as e:
        logger.error(f"Failed to read input file: {e}")
        sys.exit(1)

    required_cols = ["eprints_type", "publisher"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        logger.error(f"Missing required columns: {missing}")
        sys.exit(1)

    logger.info("Applying hard-coded canonical mapping...")
    df["publisher"] = df["publisher"].apply(apply_hardcoded_canonical)

    allowed_types = ["article", "conference_item"]
    df_filtered = df[df["eprints_type"].isin(allowed_types)].copy()
    logger.info(f"Filtered to {len(df_filtered)} relevant rows.")

    pubs_series = df_filtered["publisher"].dropna()
    if pubs_series.empty:
        logger.warning("No publisher data found after filtering. Saving unchanged file.")
        wb = load_workbook(input_file)
        wb.save(output_file)
        return

    freq_counter = Counter(pubs_series)
    unique_pubs = list(freq_counter.keys())
    pub_to_idx = {p: i for i, p in enumerate(unique_pubs)}
    n = len(unique_pubs)
    logger.info(f"Unique publishers: {n}")

    overrides = load_overrides(override_file)

    grouped_indices = defaultdict(list)
    for idx, pub in enumerate(unique_pubs):
        grouped_indices[get_group_key(pub)].append(idx)

    auto_edges = []
    review_pairs = []
    keep_pairs = set()

    override_merge_pairs = set()
    for (name1, name2), action in overrides.items():
        if name1 not in pub_to_idx or name2 not in pub_to_idx:
            continue
        i = pub_to_idx[name1]
        j = pub_to_idx[name2]
        if action is None:
            keep_pairs.add((i, j))
            keep_pairs.add((j, i))
        else:
            override_merge_pairs.add((i, j))
            override_merge_pairs.add((j, i))

    for group_key, indices in grouped_indices.items():
        if len(indices) < 2:
            continue

        logger.info(f"Comparing group '{group_key}' with {len(indices)} names")

        for a in range(len(indices)):
            for b in range(a + 1, len(indices)):
                idx_i = indices[a]
                idx_j = indices[b]
                name_i = unique_pubs[idx_i]
                name_j = unique_pubs[idx_j]

                if (idx_i, idx_j) in keep_pairs:
                    continue

                if (idx_i, idx_j) in override_merge_pairs:
                    auto_edges.append((idx_i, idx_j))
                    logger.info(f"Override merge: '{name_i}' ↔ '{name_j}'")
                    continue

                tokens_i = preprocess_name(name_i)
                tokens_j = preprocess_name(name_j)

                joined_i = " ".join(tokens_i)
                joined_j = " ".join(tokens_j)

                sim_ratio = fuzz.ratio(joined_i, joined_j)
                sim_token_sort = fuzz.token_sort_ratio(joined_i, joined_j)
                sim = max(sim_ratio, sim_token_sort)

                first_token_match = bool(tokens_i and tokens_j and tokens_i[0] == tokens_j[0])

                freq_i = freq_counter[name_i]
                freq_j = freq_counter[name_j]
                freq_ratio = max(freq_i, freq_j) / min(freq_i, freq_j) if min(freq_i, freq_j) else 0

                major_overlap = bool(set(tokens_i) & set(tokens_j) & MAJOR_PUBLISHER_TOKENS)

                if sim >= SIMILARITY_THRESHOLD_AUTO:
                    auto_edges.append((idx_i, idx_j))
                elif sim >= AUTO_BOOST_THRESHOLD and (
                    freq_ratio >= FREQ_RATIO_FOR_AUTO or first_token_match or major_overlap
                ):
                    auto_edges.append((idx_i, idx_j))
                elif sim >= SIMILARITY_THRESHOLD_REVIEW:
                    review_pairs.append((
                        idx_i, idx_j, sim, name_i, name_j,
                        freq_i, freq_j, first_token_match, major_overlap
                    ))

    logger.info(f"Auto-edges: {len(auto_edges)}")
    logger.info(f"Review pairs: {len(review_pairs)}")

    component_id_for_idx = build_components(auto_edges, n)
    comp_members = defaultdict(list)
    for idx, comp_id in component_id_for_idx.items():
        comp_members[comp_id].append(idx)

    corrections = {}
    cluster_rows = []

    for comp_id, members in comp_members.items():
        variants = [unique_pubs[idx] for idx in members]
        canonical = choose_component_canonical(variants, freq_counter)

        for variant in variants:
            corrections[variant] = canonical

        cluster_leader = min(variants, key=lambda v: (len(str(v)), v))
        canonical_freq = freq_counter[canonical]

        for idx in members:
            name = unique_pubs[idx]
            score_to_leader = fuzz.token_sort_ratio(
                " ".join(preprocess_name(name)),
                " ".join(preprocess_name(cluster_leader))
            )
            cluster_rows.append({
                "cluster_id": comp_id,
                "cluster_leader": cluster_leader,
                "canonical_name": canonical,
                "name": name,
                "frequency": freq_counter[name],
                "score_to_leader": score_to_leader,
                "cluster_size": len(variants),
                "canonical_frequency": canonical_freq,
                "all_variants": ", ".join(sorted(variants)),
            })

    cluster_df = pd.DataFrame(cluster_rows)

    review_rows = []
    for i, j, sim, name1, name2, f1, f2, first_match, major_boost in review_pairs:
        review_rows.append({
            "pub_index_1": i,
            "pub_index_2": j,
            "publisher_1": name1,
            "publisher_2": name2,
            "similarity": round(sim, 2),
            "freq_1": f1,
            "freq_2": f2,
            "first_token_match": first_match,
            "major_publisher_boost": major_boost,
            "tokens_1": " ".join(preprocess_name(name1)),
            "tokens_2": " ".join(preprocess_name(name2)),
            "suggested_action": "review",
        })

    review_df = pd.DataFrame(review_rows)
    if not review_df.empty:
        review_df.sort_values(["similarity", "freq_1", "freq_2"], ascending=[False, False, False], inplace=True)

    logger.info("Applying corrections to full dataframe...")
    df["publisher_standardised"] = df["publisher"].map(lambda x: corrections.get(x, x))

    logger.info("Saving cleaned workbook with preserved formatting...")
    try:
        wb = load_workbook(input_file, keep_links=False)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        if "publisher_standardised" in headers:
            col_out = headers.index("publisher_standardised") + 1
        else:
            col_out = len(headers) + 1
            ws.cell(row=1, column=col_out, value="publisher_standardised")

        pub_col = headers.index("publisher") + 1
        for df_idx, row in df.iterrows():
            excel_row = df_idx + 2
            ws.cell(row=excel_row, column=pub_col, value=row["publisher_standardised"])
            if col_out > len(headers):
                ws.cell(row=excel_row, column=col_out, value=row["publisher_standardised"])

        wb.save(output_file)
        logger.info(f"Cleaned file saved: {output_file}")
    except Exception as e:
        logger.error(f"Failed to save cleaned workbook: {e}")
        sys.exit(1)

    try:
        review_df.to_excel(review_file, index=False)
        logger.info(f"Review file saved: {review_file}")
    except Exception as e:
        logger.error(f"Failed to save review file: {e}")

    try:
        cluster_df.to_csv(cluster_file, index=False)
        logger.info(f"Cluster summary saved: {cluster_file}")
    except Exception as e:
        logger.error(f"Failed to save cluster file: {e}")

    logger.info("=" * 70)
    logger.info("SUMMARY")
    logger.info(f"Rows processed: {len(df)}")
    logger.info(f"Unique publishers before: {len(unique_pubs)}")
    logger.info(f"Clusters formed: {len([m for m in comp_members.values() if len(m) > 1])}")
    logger.info(f"Review pairs: {len(review_pairs)}")
    logger.info("=" * 70)

# ================================================================
# 6. CLI entry point
# ================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Robust publisher name standardisation pipeline."
    )
    parser.add_argument("--input", default="inputs/metadata_extract_20260127.xlsx")
    parser.add_argument("--output", default="outputs/metadata_extract_20260127_publishers_cleaned.xlsx")
    parser.add_argument("--review", default="outputs/publisher_review_index.xlsx")
    parser.add_argument("--clusters", default="outputs/publisher_cluster_summary.csv")
    parser.add_argument("--override", default=None)
    args = parser.parse_args()

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    Path(args.review).parent.mkdir(parents=True, exist_ok=True)
    Path(args.clusters).parent.mkdir(parents=True, exist_ok=True)

    process_publishers(args.input, args.output, args.review, args.clusters, args.override)