#!/usr/bin/env python3
"""
Task 3 – Author Name Standardisation

Analyses first_name and last_name columns, using the staff identifier (id) where
available to assist in duplicate detection and identity resolution.

Generates:
  - A cleaned authors file with standardised names
  - A review index for manual validation of uncertain matches
  - A cluster summary
  - A duplicate candidates file for names with different IDs

Usage:
    python author_name_standardisation_task_three_activity_one.py --input authors_20260127_WorkingFile.xlsx --output authors_cleaned.xlsx
"""

import argparse
import logging
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from rapidfuzz import fuzz

# ================================================================
# 1. Logging
# ================================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("author_standardisation.log", encoding="utf-8"),
    ],
)
logger = logging.getLogger(__name__)

# ================================================================
# 2. Configuration
# ================================================================
SIMILARITY_THRESHOLD_AUTO = 85
SIMILARITY_THRESHOLD_REVIEW = 75

TITLE_STOPWORDS = {
    "dr",
    "prof",
    "professor",
    "mr",
    "mrs",
    "ms",
    "miss",
    "sir",
    "dame",
    "lord",
    "lady",
    "rev",
    "revd",
    "canon",
    "fr",
    "brother",
    "sister",
}

SUFFIX_STOPWORDS = {
    "jr",
    "sr",
    "i",
    "ii",
    "iii",
    "iv",
    "v",
    "vi",
    "vii",
    "viii",
    "ix",
    "x",
    "phd",
    "md",
    "dphil",
    "mbbs",
    "frcp",
    "frs",
}

# ================================================================
# 3. Helper Functions
# ================================================================


def clean_name_part(text):
    """Clean a single name part (first_name or last_name)."""
    if pd.isna(text) or text == "":
        return ""
    s = str(text).strip().lower()
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def title_case_name(text):
    """Convert to title case, handling hyphenated names."""
    if pd.isna(text) or text == "":
        return ""
    return " ".join(part.title() for part in str(text).split())


def preprocess_full_name(first_name, last_name):
    """
    Combine first_name and last_name into a single string for comparison.
    Returns full name in "last, first" format (or just last if first missing).
    Also returns a normalised version without titles/suffixes.
    """
    first = clean_name_part(first_name)
    last = clean_name_part(last_name)

    # Build full name as "last, first" if we have both
    if first and last:
        full = f"{last}, {first}"
    elif last:
        full = last
    elif first:
        full = first
    else:
        full = ""

    # Remove titles and suffixes for matching
    tokens = full.split()
    tokens = [t for t in tokens if t not in TITLE_STOPWORDS]
    tokens = [t for t in tokens if t not in SUFFIX_STOPWORDS]
    normalised = " ".join(tokens)

    return {
        "full": full,
        "first": first,
        "last": last,
        "normalised": normalised,
    }


def get_group_key(first_name, last_name):
    """Generate a grouping key based on the last name."""
    first = clean_name_part(first_name)
    last = clean_name_part(last_name)
    if last:
        return last[:4]
    elif first:
        return first[:4]
    else:
        return "UNKN"


def build_components(edges, num_items):
    """Union-Find algorithm to group indices into connected components."""
    parent = list(range(num_items))

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(x, y):
        rx, ry = find(x), find(y)
        if rx != ry:
            parent[ry] = rx

    for i, j in edges:
        union(i, j)

    comp_map = {}
    for i in range(num_items):
        root = find(i)
        comp_map.setdefault(root, []).append(i)

    comp_id_for_item = {}
    for comp_id, members in enumerate(comp_map.values(), start=1):
        for idx in members:
            comp_id_for_item[idx] = comp_id
    return comp_id_for_item


def load_overrides(override_file):
    """Load manual override file (CSV/Excel)."""
    if not override_file:
        return {}
    override_path = Path(override_file)
    if not override_path.exists():
        return {}

    logger.info(f"Loading overrides from {override_file}")
    try:
        if override_path.suffix.lower() == ".csv":
            ov = pd.read_csv(override_file)
        else:
            ov = pd.read_excel(override_file)
    except Exception as e:
        logger.error(f"Failed to read override file: {e}")
        return {}

    overrides = {}
    for _, row in ov.iterrows():
        if "id_1" in ov.columns and "id_2" in ov.columns:
            key1 = str(row["id_1"])
            key2 = str(row["id_2"])
        elif "resource_id_1" in ov.columns and "resource_id_2" in ov.columns:
            key1 = str(row["resource_id_1"])
            key2 = str(row["resource_id_2"])
        else:
            key1 = str(row["name_1"])
            key2 = str(row["name_2"])

        action = str(row["action"]).strip().lower()
        if action == "keep":
            overrides[(key1, key2)] = None
        else:
            overrides[(key1, key2)] = action
        overrides[(key2, key1)] = overrides[(key1, key2)]

    logger.info(f"Loaded {len(overrides) // 2} override pairs.")
    return overrides


# ================================================================
# 4. Main Processing Function
# ================================================================


def process_authors(
    input_file, output_file, review_file, cluster_file, override_file=None
):
    logger.info("=" * 70)
    logger.info("Author Name Standardisation Pipeline")
    logger.info("=" * 70)
    logger.info(f"Input: {input_file}")
    logger.info(f"Output: {output_file}")

    input_path = Path(input_file)
    if not input_path.exists():
        logger.error(f"Input file not found: {input_file}")
        sys.exit(1)

    # ------------------------------------------------------------
    # 4a. Read the data
    # ------------------------------------------------------------
    try:
        df = pd.read_excel(input_file, engine="openpyxl")
    except Exception as e:
        logger.error(f"Failed to read input file: {e}")
        sys.exit(1)

    logger.info(f"Loaded {len(df)} rows")

    # Check required columns
    required_cols = ["first_name", "last_name"]
    missing = [c for c in required_cols if c not in df.columns]
    if missing:
        logger.error(f"Missing required columns: {missing}")
        sys.exit(1)

    has_id = "id" in df.columns
    has_resource_id = "resource_id" in df.columns

    # ------------------------------------------------------------
    # 4b. Preprocess names
    # ------------------------------------------------------------
    logger.info("Preprocessing names...")

    df["first_name_clean"] = df["first_name"].apply(clean_name_part)
    df["last_name_clean"] = df["last_name"].apply(clean_name_part)

    name_data = df.apply(
        lambda row: preprocess_full_name(row["first_name"], row["last_name"]), axis=1
    )

    df["full_name"] = [d["full"] for d in name_data]
    df["normalised_name"] = [d["normalised"] for d in name_data]
    df["first_name_normalised"] = [d["first"] for d in name_data]
    df["last_name_normalised"] = [d["last"] for d in name_data]

    # Create a unique record identifier
    if has_resource_id:
        df["record_id"] = df["resource_id"].astype(str)
    elif has_id:
        df["record_id"] = df["id"].astype(str)
    else:
        df["record_id"] = df.index.astype(str)

    # ------------------------------------------------------------
    # 4c. Group using ID where available (strongest signal)
    # ------------------------------------------------------------
    auto_edges = []
    review_pairs = []
    id_groups = defaultdict(list)

    if has_id:
        logger.info("Using 'id' column for initial grouping...")
        for idx, row in df.iterrows():
            if pd.notna(row["id"]) and row["id"] != "":
                id_groups[str(row["id"])].append(idx)

        for id_val, indices in id_groups.items():
            if len(indices) > 1:
                for i in range(len(indices)):
                    for j in range(i + 1, len(indices)):
                        auto_edges.append((indices[i], indices[j]))
        logger.info(f"Auto-merged {len(auto_edges)} pairs using ID")

    # ------------------------------------------------------------
    # 4d. Fuzzy matching for remaining records
    # ------------------------------------------------------------
    logger.info("Applying fuzzy matching...")

    grouped_indices = defaultdict(list)
    for idx, row in df.iterrows():
        key = get_group_key(row["first_name"], row["last_name"])
        grouped_indices[key].append(idx)

    already_merged = set(auto_edges)

    for group_key, indices in grouped_indices.items():
        if len(indices) < 2:
            continue

        for a in range(len(indices)):
            for b in range(a + 1, len(indices)):
                idx_i = indices[a]
                idx_j = indices[b]

                if (idx_i, idx_j) in already_merged or (idx_j, idx_i) in already_merged:
                    continue

                name_i = str(df.loc[idx_i, "normalised_name"])
                name_j = str(df.loc[idx_j, "normalised_name"])

                if not name_i or not name_j:
                    continue

                sim = fuzz.token_sort_ratio(name_i, name_j)

                if sim >= SIMILARITY_THRESHOLD_AUTO:
                    auto_edges.append((idx_i, idx_j))
                elif sim >= SIMILARITY_THRESHOLD_REVIEW:
                    review_pairs.append(
                        (
                            idx_i,
                            idx_j,
                            sim,
                            str(df.loc[idx_i, "full_name"]),
                            str(df.loc[idx_j, "full_name"]),
                            str(df.loc[idx_i, "record_id"]),
                            str(df.loc[idx_j, "record_id"]),
                            str(df.loc[idx_i, "id"]) if has_id else None,
                            str(df.loc[idx_j, "id"]) if has_id else None,
                            df.loc[idx_i, "resource_id"] if has_resource_id else None,
                            df.loc[idx_j, "resource_id"] if has_resource_id else None,
                        )
                    )

    logger.info(f"Auto-edges: {len(auto_edges)}")
    logger.info(f"Review pairs: {len(review_pairs)}")

    # ------------------------------------------------------------
    # 4e. Build components and choose canonical names
    # ------------------------------------------------------------
    n = len(df)
    component_id_for_idx = build_components(auto_edges, n)
    comp_members = defaultdict(list)

    for idx, comp_id in component_id_for_idx.items():
        comp_members[comp_id].append(idx)

    corrections = {}
    cluster_rows = []

    for comp_id, members in comp_members.items():
        if len(members) < 2:
            for idx in members:
                record_id = str(df.loc[idx, "record_id"])
                # Use the full_name as canonical, but title-case it
                full = str(df.loc[idx, "full_name"])
                # Try to split into last and first for title-casing
                if ", " in full:
                    last, first = full.split(", ", 1)
                    canonical = f"{title_case_name(last)}, {title_case_name(first)}"
                else:
                    canonical = title_case_name(full)
                corrections[record_id] = canonical
            continue

        variants = [str(df.loc[idx, "full_name"]) for idx in members]
        variant_counts = Counter(variants)
        canonical_raw = max(
            variant_counts, key=lambda v: (variant_counts[v], len(str(v)))
        )

        # Title-case the canonical name
        if ", " in canonical_raw:
            last, first = canonical_raw.split(", ", 1)
            canonical = f"{title_case_name(last)}, {title_case_name(first)}"
        else:
            canonical = title_case_name(canonical_raw)

        for idx in members:
            record_id = str(df.loc[idx, "record_id"])
            corrections[record_id] = canonical

        cluster_rows.append(
            {
                "cluster_id": comp_id,
                "cluster_size": len(members),
                "canonical_name": canonical,
                "member_ids": ", ".join(
                    [str(df.loc[idx, "record_id"]) for idx in members]
                ),
                "member_names": ", ".join(
                    [str(df.loc[idx, "full_name"]) for idx in members]
                ),
            }
        )

    # ------------------------------------------------------------
    # 4f. Apply corrections and split into first/last
    # ------------------------------------------------------------
    df["author_standardised"] = df["record_id"].map(corrections)

    # For any unmatched records, use the original full name (title-cased)
    def title_case_full(name):
        if pd.isna(name) or name == "":
            return ""
        if ", " in name:
            last, first = name.split(", ", 1)
            return f"{title_case_name(last)}, {title_case_name(first)}"
        else:
            return title_case_name(name)

    df["author_standardised"] = df["author_standardised"].apply(
        lambda x: title_case_full(x) if pd.notna(x) else ""
    )

    # Split into first and last names (title-cased already)
    def split_name(name):
        if pd.isna(name) or name == "":
            return "", ""
        if ", " in name:
            parts = name.split(", ")
            return parts[1] if len(parts) > 1 else "", parts[0]
        else:
            # If no comma, assume it's just last name
            return "", name

    df[["first_name_standardised", "last_name_standardised"]] = df[
        "author_standardised"
    ].apply(lambda x: pd.Series(split_name(x)))

    # ------------------------------------------------------------
    # 4g. Save the cleaned workbook (fixed column handling)
    # ------------------------------------------------------------
    logger.info("Saving cleaned workbook...")

    try:
        wb = load_workbook(input_file, keep_links=False, data_only=True)

        # --- Aggressive cleanup ---
        connections = getattr(wb, "connections", None)
        if connections is not None:
            if hasattr(connections, "_connections"):
                connections._connections.clear()
            setattr(wb, "connections", None)
        if getattr(wb, "external_links", None) is not None:
            setattr(wb, "external_links", [])
        if hasattr(wb, "_external_links"):
            setattr(wb, "_external_links", [])
        if hasattr(wb, "_connections"):
            setattr(wb, "_connections", [])
        if hasattr(wb, "_tables"):
            setattr(wb, "_tables", {})
        current_ranges = getattr(wb, "named_ranges", None)
        if current_ranges:
            filtered = [
                nr for nr in current_ranges if not getattr(nr, "external", False)
            ]
            setattr(wb, "named_ranges", filtered)

        ws = wb.active
        if ws is None:
            logger.error("Active worksheet is None – cannot proceed.")
            sys.exit(1)

        # Get existing headers
        headers = [cell.value for cell in ws[1]]

        # Add new columns (track their indices)
        new_cols = [
            "first_name_standardised",
            "last_name_standardised",
            "author_standardised",
        ]
        col_map = {}  # col_name -> column index (1-based)

        for col_name in new_cols:
            if col_name in headers:
                col_map[col_name] = headers.index(col_name) + 1
            else:
                # Add new column
                new_col_idx = len(headers) + 1
                ws.cell(row=1, column=new_col_idx, value=col_name)
                col_map[col_name] = new_col_idx
                headers.append(col_name)  # update headers

        # Write data rows
        for df_idx, row in df.iterrows():
            excel_row = df_idx + 2
            for col_name in new_cols:
                if col_name in df.columns:
                    ws.cell(
                        row=excel_row,
                        column=col_map[col_name],
                        value=str(row[col_name]),
                    )

        wb.save(output_file)
        logger.info(f"Cleaned file saved: {output_file}")

    except Exception as e:
        logger.error(f"Failed to save output file: {e}")
        sys.exit(1)

    # ------------------------------------------------------------
    # 4h. Generate review index with extra columns
    # ------------------------------------------------------------
    review_rows = []
    for (
        idx_i,
        idx_j,
        sim,
        name1,
        name2,
        rec_id1,
        rec_id2,
        id1,
        id2,
        res_id1,
        res_id2,
    ) in review_pairs:
        # Calculate frequency for each (we have full df)
        freq1 = len(df[df["record_id"] == rec_id1])  # could be multiple rows
        freq2 = len(df[df["record_id"] == rec_id2])
        same_id_flag = id1 == id2 and id1 is not None and id1 != ""

        review_rows.append(
            {
                "record_index_1": idx_i,
                "record_index_2": idx_j,
                "resource_id_1": res_id1,
                "resource_id_2": res_id2,
                "record_id_1": rec_id1,
                "record_id_2": rec_id2,
                "id_1": id1,
                "id_2": id2,
                "name_1": name1,
                "name_2": name2,
                "similarity": round(sim, 2),
                "freq_1": freq1,
                "freq_2": freq2,
                "same_id": same_id_flag,
                "suggested_action": "Merge" if same_id_flag else "Review",
            }
        )

    review_df = pd.DataFrame(review_rows)
    if not review_df.empty:
        # Sort by similarity descending, but put same_id at top
        review_df.sort_values(
            ["same_id", "similarity"], ascending=[False, False], inplace=True
        )

    try:
        review_df.to_excel(review_file, index=False)
        logger.info(f"Review file saved: {review_file}")
    except Exception as e:
        logger.error(f"Failed to save review file: {e}")

    # ------------------------------------------------------------
    # 4i. Save cluster summary
    # ------------------------------------------------------------
    cluster_df = pd.DataFrame(cluster_rows)
    try:
        cluster_df.to_csv(cluster_file, index=False)
        logger.info(f"Cluster summary saved: {cluster_file}")
    except Exception as e:
        logger.error(f"Failed to save cluster file: {e}")

    # ------------------------------------------------------------
    # 4j. Summary
    # ------------------------------------------------------------
    logger.info("=" * 70)
    logger.info("SUMMARY")
    logger.info(f"Total rows processed: {len(df)}")
    logger.info(f"Unique IDs used for grouping: {len(id_groups) if has_id else 0}")
    logger.info(
        f"Clusters formed: {len([m for m in comp_members.values() if len(m) > 1])}"
    )
    logger.info(f"Review pairs: {len(review_pairs)}")
    logger.info("=" * 70)


# ================================================================
# 5. Command-line Entry Point
# ================================================================
if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Author name standardisation pipeline."
    )
    parser.add_argument(
        "--input",
        default="authors_20260127_WorkingFile.xlsx",
        help="Path to input authors file",
    )
    parser.add_argument(
        "--output",
        default="Outputs/authors_cleaned.xlsx",
        help="Path to output cleaned file",
    )
    parser.add_argument(
        "--review",
        default="Outputs/author_review_index.xlsx",
        help="Path to output review index",
    )
    parser.add_argument(
        "--clusters",
        default="Outputs/author_cluster_summary.csv",
        help="Path to output cluster summary",
    )
    parser.add_argument(
        "--override", default=None, help="Path to manual override file (optional)"
    )

    args = parser.parse_args()

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    Path(args.review).parent.mkdir(parents=True, exist_ok=True)
    Path(args.clusters).parent.mkdir(parents=True, exist_ok=True)

    process_authors(args.input, args.output, args.review, args.clusters, args.override)
